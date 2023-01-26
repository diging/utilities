from openpyxl import load_workbook, Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Alignment, Border, Side
from datetime import time, datetime, date, timedelta
from openpyxl.utils import get_column_letter
from schedule_utils import handle_lunch, format_cell, set_border



def main():
    hours_wb = load_workbook(filename = '2023_Spring_Semester_Schedule1.xlsx')
    hours_sheet = hours_wb['Sheet1']

    # declare styles here for easy modification
    header = Font(name='Calibri', size=20, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='00000000')
    days = Font(name='Calibri', size=15, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='00000000')
    cell_aligment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
    time_borders = Border(left=Side(border_style='medium', color='00C0C0C0'), right=Side(border_style='medium', color='00C0C0C0'),  bottom=Side(border_style='medium', color='00C0C0C0'), top=Side(border_style='medium', color='00C0C0C0'))
    rectangle_border = Border(left=Side(border_style='medium', color='00000000'), right=Side(border_style='medium', color='00000000'),  bottom=Side(border_style='medium', color='00000000'), top=Side(border_style='medium', color='00000000'))

    new_wb = Workbook()
    new_sheet = new_wb.active
    hours_col_a = hours_sheet['A']

    new_sheet['A1'] = 'DigInG Schedule'
    new_sheet['A1'].font = header

    new_sheet.column_dimensions["A"].width = 13

    for x in range(2, 20):
        new_sheet.column_dimensions[get_column_letter(x)].width = 8

    row_number = 3

    days_of_weeks_strings = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

    background_colors = ["FFBA84", '90B44B', '7DB9DE', '787D7B', '005CAF', 'EFBB24', '994639', '1B813E', '52433D', '81C7D4', '8E354A', '554236', 'F7D94C', '646A58', '7B90D2']

    worker_map = {}

    for day in days_of_weeks_strings:
        new_sheet.cell(row=row_number, column=1, value=day).font = days
        count = 0
        start_column = 2
        end_column = start_column + 1
        time_list = ['8:00', '9:00', '10:00', '11:00', '12:00', '1:00', '2:00', '3:00', '4:00']
        while count < 9:
            new_sheet.cell(row=row_number, column=start_column, value=time_list[count]).alignment = cell_aligment
            new_sheet.cell(row=row_number, column=start_column, value=time_list[count]).border = time_borders
            new_sheet.merge_cells(start_row=row_number, start_column=start_column, end_row=row_number, end_column=end_column)
            start_column = end_column + 1
            end_column = start_column + 1
            count += 1
        row_number += 1
        for index, name in enumerate(hours_col_a):
            if index != 0:
                new_sheet['A' + str(row_number)] = name.value
                if name.value in worker_map:
                    worker_map[name.value].append(row_number)
                else:
                    worker_map[name.value] = [row_number]
                row_number += 1
        row_number += 1
    total_scheduled_hours = {}

    num_workers = len(worker_map) - 1
    col_start = 2
    col_end = 19
    row_start = 3
    row_end = row_start + num_workers
    for day in days_of_weeks_strings:
        set_border(new_sheet, f'{new_sheet.cell(row_start, col_start).coordinate}:{new_sheet.cell(row_end, col_end).coordinate}', background_colors)
        row_start += num_workers + 7
        row_end += num_workers + 7

    for index, row in enumerate(hours_sheet.iter_rows()):
        # check if there is a name, if not consider it the end of the list
        if not row[0].value:
            break
        # skip header and only run if max ours isnt 0
        if index != 0 and int(row[1].value) != 0:
            name = row[0].value
            total_scheduled_hours[name] = timedelta(minutes=0)
            print('^^^^^^^^^^^^^^^^^^^^^^^^^' + name + '^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^')
            max_avail_minutes = int(row[1].value) * 60
            scheduled_minutes = 0
            row_numbers = worker_map[name]
            day = 1
            start = 2
            end = 3
            while day < 6:
                print(f"******************Day: {day}*****************")
                start_string = row[start].value
                end_string = row[end].value
                if isinstance(start_string, time) and isinstance(end_string, time):
                    #change var name not called string - its a time object
                    print(f"Start String: {start_string}, End String: {end_string}")
                    start_string_2 = row[start + 2].value
                    end_string_2 = row[end + 2].value
                    print(f"start 2: {start_string_2}, end 2: {end_string_2}")
                    start_column = 2 + ((start_string.hour - 8)*2) + (start_string.minute / 30)
                    print(f"Start Column: {start_column}")
                    duration = datetime.combine(date.min, end_string) - datetime.combine(date.min, start_string)
                    total_scheduled_hours[name] += duration
                    print('Duration: {}'.format(duration ))
                    total_cells = duration.seconds / 60 / 30
                    temp_scheduled_minutes = scheduled_minutes
                    temp_scheduled_minutes += duration.seconds / 60
                    if temp_scheduled_minutes > max_avail_minutes:
                        left_mins = max_avail_minutes - scheduled_minutes
                        # only schedule if shift is longer then 120 mins
                        # if left_mins >= 120:
                        #     print("Mins left: {}".format(type(left_mins)))
                        #     end_string = (datetime.combine(date.today(), start_string) + timedelta(minutes=left_mins)).time()
                        #     total_cells = left_mins / 15
                        #     total_scheduled_hours[name] -= timedelta(minutes=30)
                    # current times go one cell past everytime
                    end_column = (start_column + total_cells) - 1
                    schedule_cell = format_cell(new_sheet, row_numbers[day-1], start_column, end_column, background_colors[index], rectangle_border)
                    cell_string = datetime.strptime(start_string.isoformat(timespec='minutes'), "%H:%M").strftime("%I:%M %p") +' - ' + datetime.strptime(end_string.isoformat(timespec='minutes'), "%H:%M").strftime("%I:%M %p")
                    total_scheduled_hours[name], schedule_cell.value = handle_lunch(duration, cell_string, total_scheduled_hours[name])
                    new_sheet.cell(row=row_numbers[day-1], column=start_column).alignment = cell_aligment
                    # check for break and second daily shift
                    if isinstance(start_string_2, time):
                        duration_2 = datetime.combine(date.min, end_string_2) - datetime.combine(date.min, start_string_2)
                        break_duration = datetime.combine(date.min, start_string_2) - datetime.combine(date.min, end_string)
                        print(f"break duration: {break_duration}")
                        break_quarter_increments = divmod((break_duration.total_seconds()/60), 30)
                        duration_2_quarter_increments = divmod((duration_2.total_seconds()/60), 30)
                        print(f"duration increments: {duration_2_quarter_increments}")

                        total_scheduled_hours[name] += duration_2
                        print('duration_2: {}'.format(duration_2 ))
                        total_cells = duration_2.seconds / 60 / 30
                        temp_scheduled_minutes = scheduled_minutes
                        temp_scheduled_minutes += duration_2.seconds / 60
                        if temp_scheduled_minutes > max_avail_minutes:
                            left_mins = max_avail_minutes - scheduled_minutes
                            # only schedule if shift is longer then 120 mins
                            # if left_mins >= 120:
                            #     print("Mins left: {}".format(type(left_mins)))
                            #     end_string = (datetime.combine(date.today(), start_string) + timedelta(minutes=left_mins)).time()
                            #     total_cells = left_mins / 15
                            #     total_scheduled_hours[name] -= timedelta(minutes=30)
                        print(f"Incoming Start: {start_column}, end column: {end_column}")
                        start_column = (end_column + break_quarter_increments[0]) + 1
                        end_column = (start_column + duration_2_quarter_increments[0]) - 1
                        print(f"After Start: {start_column}, end column: {end_column}")
                        schedule_cell = format_cell(new_sheet, row_numbers[day-1], start_column, end_column, background_colors[index], rectangle_border)
                        cell_string = datetime.strptime(start_string_2.isoformat(timespec='minutes'), "%H:%M").strftime("%I:%M %p") +' - ' + datetime.strptime(end_string_2.isoformat(timespec='minutes'), "%H:%M").strftime("%I:%M %p")
                        total_scheduled_hours[name], schedule_cell.value = handle_lunch(duration_2, cell_string, total_scheduled_hours[name])
                        new_sheet.cell(row=row_numbers[day-1], column=start_column).alignment = cell_aligment
                    scheduled_minutes = temp_scheduled_minutes
                start = start + 4
                end = end + 4
                day += 1
        print(total_scheduled_hours)
    print(row_number)
    for k, v in total_scheduled_hours.items():
        new_sheet.cell(row=row_number, column=1).value = k + ': ' + str(v.total_seconds() / 3600)
        row_number += 1



    new_wb.save(filename = 'new_schedule.xlsx')





if __name__ == '__main__':
    main()