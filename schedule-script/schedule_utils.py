from datetime import time, datetime, date, timedelta
from openpyxl.styles import Font, Color, PatternFill, Alignment, Border, Side

def handle_lunch(duration, cell_string, total_scheduled_hours):
    if (duration.seconds / 60) > 300:
        print('duration_2: {}'.format(duration.seconds))
        cell_string = cell_string + ' (1/2 hour lunch)'
        total_scheduled_hours -= timedelta(minutes=30)
    return total_scheduled_hours, cell_string

def format_cell(new_sheet, row_number, start_column, end_column, background_colors, rectangle_border):
    schedule_cell = new_sheet.cell(row=row_number, column=start_column)
    schedule_cell.fill = PatternFill("solid", fgColor=background_colors)
    schedule_cell.border = rectangle_border
    new_sheet.merge_cells(start_row=row_number, start_column=start_column, end_row=row_number, end_column=end_column)
    return schedule_cell

def set_border(new_sheet, cell_range, background_colors):
    thick = Side(border_style="thick", color="000000")
    thin = Side(border_style="thin", color="000000")
    for row in new_sheet[cell_range]:
        for c in row:
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for c in new_sheet[cell_range][0][1:-1]:
        c.border = Border(top=thick, right=thin)
    for c in new_sheet[cell_range][-1][1:-1]:
        c.border = Border(bottom=thick, right=thin)
    for index, row in enumerate(new_sheet[cell_range]):
        row[0].border = Border(left=Side(border_style="thick", color=background_colors[index]), bottom= thin, top=thin)
        row[-1].border = Border(right=thick, bottom=thin, top=thin)
    new_sheet[cell_range][0][0].border = Border(left=thick, top=thick)
    new_sheet[cell_range][0][-1].border = Border(right=thick, top=thick)
    new_sheet[cell_range][-1][0].border = Border(left=Side(border_style="thick", color=background_colors[12]), bottom=thick)
    new_sheet[cell_range][-1][-1].border = Border(right=thick, bottom=thick)
